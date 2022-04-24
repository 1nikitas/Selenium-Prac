import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import datetime
from config import email, passw


book = openpyxl.Workbook()
sheet = book.active

sheet.cell(row=1, column=1).value = "Направление"
sheet.cell(row=1, column=2).value = "Тип Фургона"
sheet.cell(row=1, column=3).value = "Вес"
sheet.cell(row=1, column=4).value = "Объем"
sheet.cell(row=1, column=5).value = "Тип Транспорта"
sheet.cell(row=1, column=6).value = "Габариты транспорта"
sheet.cell(row=1, column=7).value = "Тип погрузки в транспорт"
sheet.cell(row=1, column=8).value = "От куда"
sheet.cell(row=1, column=9).value = "Куда"
sheet.cell(row=1, column=10).value = "Дата погрузки"
sheet.cell(row=1, column=11).value = "Ставка"
sheet.cell(row=1, column=12).value = "Контакты"
sheet.cell(row=1, column=13).value = "Коментарий"
sheet.cell(row=1, column=14).value = "Компания"
sheet.cell(row=1, column=15).value = "Код"
sheet.cell(row=1, column=16).value = "Перевозчик"
sheet.cell(row=1, column=17).value = "р"

time1 = datetime.datetime.now()
browser = webdriver.Chrome('chromedriver.exe')
browser.get("https://id.ati.su/login/?next=https%3A%2F%2Fati.su%2F")

login = browser.find_element(By.ID, 'login')
login.send_keys(email)
password = browser.find_element(By.ID, 'password')
password.send_keys(passw)
button_submit = browser.find_element(By.ID, 'action-login')
button_submit.click()

time.sleep(3)

browser.get("https://trucks.ati.su/?Weight=1")

time.sleep(3)

max_number = int(browser.find_element(By.CLASS_NAME, "total-index_CpiGH").text.replace(" ", ""))
# dont forget max_number + 1

row_n1 = 2

for i in range(4):

    time.sleep(3)
    blocks = browser.find_elements(By.CLASS_NAME, value="sc-XxNYO")
    for block in blocks:

        firm_params = block.find_elements(By.CLASS_NAME,
                                          value="firm-params")  # Бондаренко Андрей Александрович, ИП Код:240846, Перевозчик, Sochi
        for firm_param in firm_params:
            info = firm_param.text.replace(",", "").split("\n")
            try:
                sheet.cell(row=row_n1, column=14).value = info[0]
            except IndexError:
                sheet.cell(row=row_n1, column=14).value = ""

            try:
                sheet.cell(row=row_n1, column=15).value = info[1]
            except IndexError:
                sheet.cell(row=row_n1, column=15).value = ""
            try:
                sheet.cell(row=row_n1, column=16).value = info[2]
            except:
                sheet.cell(row=row_n1, column=16).value = ""
            try:
                sheet.cell(row=row_n1, column=17).value = info[3]
            except IndexError:
                sheet.cell(row=row_n1, column=17).value = ""

        phone_number_amount = len(block.find_elements(By.CLASS_NAME, value="sc-ctqQKy"))
        phone_num = ""

        if phone_number_amount == 0:
            phone_num = "Closed"
            sheet.cell(row=row_n1, column=12).value = phone_num
        elif phone_number_amount == 1:
            phone_numbers = block.find_elements(By.CLASS_NAME, value="sc-ctqQKy")
            for phone_number in phone_numbers:
                phone_num += phone_number.text.replace(",", " ")
            sheet.cell(row=row_n1, column=12).value = str(phone_num)
        elif phone_number_amount == 2:
            phone_numbers = block.find_elements(By.CLASS_NAME, value="sc-ctqQKy")
            for phone_number in phone_numbers:
                phone_num += phone_number.text.replace(",", " ")
            sheet.cell(row=row_n1, column=12).value = str(phone_num)
        elif phone_number_amount == 3:
            phone_numbers = block.find_elements(By.CLASS_NAME, value="sc-ctqQKy")
            for phone_number in phone_numbers:
                phone_num += phone_number.text.replace(",", " ")
            sheet.cell(row=row_n1, column=12).value = str(phone_num)
        elif phone_number_amount == 4:
            phone_numbers = block.find_elements(By.CLASS_NAME, value="sc-ctqQKy")
            for phone_number in phone_numbers:
                phone_num += phone_number.text.replace(",", " ")
            sheet.cell(row=row_n1, column=12).value = str(phone_num)
        elif phone_number_amount == 5:
            phone_numbers = block.find_elements(By.CLASS_NAME, value="sc-ctqQKy")
            for phone_number in phone_numbers:
                phone_num += phone_number.text.replace(",", " ")
            sheet.cell(row=row_n1, column=12).value = str(phone_num)
        elif phone_number_amount == 6:
            phone_numbers = block.find_elements(By.CLASS_NAME, value="sc-ctqQKy")
            for phone_number in phone_numbers:
                phone_num += phone_number.text.replace(",", " ")
            sheet.cell(row=row_n1, column=12).value = str(phone_num)

        notes = browser.find_elements(By.CLASS_NAME, value="note-text")
        for note in notes:
            info = note.text
            try:
                sheet.cell(row=row_n1, column=13).value = info
            except:
                sheet.cell(row=row_n1, column=13).value = " "

        truck_dimensions = block.find_elements(By.CSS_SELECTOR, "p[data-qa='truck-dimensions']")
        for truck_dimension in truck_dimensions:
            info = truck_dimension.text
            try:
                sheet.cell(row=row_n1, column=6).value = info
            except:
                sheet.cell(row=row_n1, column=6).value = ""

        loadings = block.find_elements(By.CSS_SELECTOR, "p[data-qa='loading-city']")
        for loading in loadings:
            info = loading.text
            sheet.cell(row=row_n1, column=8).value = info

        uploadings = block.find_elements(By.CSS_SELECTOR, "p[data-qa='main-unloading-point-name']")
        for uploading in uploadings:
            info = uploading.text
            try:
                sheet.cell(row=row_n1, column=9).value = info
            except:
                sheet.cell(row=row_n1, column=9).value = ""

        periods = block.find_elements(By.CSS_SELECTOR, "p[data-qa='loading-periodicity']")
        for period in periods:
            info = period.text
            sheet.cell(row=row_n1, column=10).value = info

        rates = block.find_elements(By.CSS_SELECTOR, "div[data-qa='rate']")
        for rate in rates:
            info = rate.text
            try:
                sheet.cell(row=row_n1, column=11).value = info
            except:
                sheet.cell(row=row_n1, column=11).value = ""

        truck_sels = block.find_elements(By.CSS_SELECTOR, "label[data-qa='checkbox']")
        for truck_sel in truck_sels:
            if truck_sel.text == "Направл.":
                continue
            else:
                info = truck_sel.text
                sheet.cell(row=row_n1, column=1).value = info

        truck_infos = block.find_elements(By.CSS_SELECTOR, "p[data-qa='truck-info']")
        for truck_info in truck_infos:
            info = truck_info.text.replace(" ", "").split(",")
            sheet.cell(row=row_n1, column=2).value = info[0]
            sheet.cell(row=row_n1, column=3).value = info[1]
            sheet.cell(row=row_n1, column=4).value = info[2]
            try:
                sheet.cell(row=row_n1, column=5).value = info[3]
            except:
                sheet.cell(row=row_n1, column=5).value = ""

        truck_loadings = block.find_elements(By.CSS_SELECTOR, "p[data-qa='truck-loading-params']")
        for truck_loading in truck_loadings:
            info = truck_loading.text
            try:
                sheet.cell(row=row_n1, column=7).value = info
            except:
                sheet.cell(row=row_n1, column=7).value = ""

        row_n1 += 1

    button_next = browser.find_element(By.CLASS_NAME, 'next_e4MPo')
    button_next.click()

book.save("Result.xlsx")
print("Файл успешно сформирован!")
book.close()